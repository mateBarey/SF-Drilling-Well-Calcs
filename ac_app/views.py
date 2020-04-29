from django.shortcuts import render, redirect
from ac_app.forms import UserForm, UserProfileInfoForm
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect,HttpResponse, HttpResponseBadRequest
from django.contrib.auth import authenticate, login, logout
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from arcpy.da import SearchCursor
from arcpy import SelectLayerByLocation_management, MakeFeatureLayer_management, Delete_management, SpatialReference, PointGeometry
from arcpy import env
import pandas as pd
import os
import math
import pyexcel
from django.template import RequestContext
import django_excel as excel
import io 
from django.core.files import File
from pyexcel.cookbook import merge_all_to_a_book
import glob
from ac_app.models import AcFlaggedWell
from pyexcel._compact import OrderedDict





# Create your views here.
def index(request):
    return render(request,'ac_app/index.html')

@login_required
def special(request):
    # Remember to also set login url in settings.py!
    # LOGIN_URL = '/basic_app/user_login/'
    return HttpResponse("You are logged in. Nice!")

@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect(reverse('index'))

def register(request):
    registered = False
    if request.method == "POST":
        user_form = UserForm(data=request.POST)
        profile_form = UserProfileInfoForm(data=request.POST)

        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            user.set_password(user.password)
            user.save()

            profile = profile_form.save(commit=False)
            profile.user = user

            if 'profile_pic' in request.FILES:
                profile.profile_pic = request.FILES['profile_pic']
            profile.save()

            registered = True
        else:
            print(user_form.errors,profile_form.errors)
    else:
        user_form= UserForm()
        profile_form = UserProfileInfoForm()

    return render(request,'ac_app/registration.html',
                            {'user_form':user_form,
                            'profile_form':profile_form,
                            'registered':registered})

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:
            if user.is_active:
                login(request,user)
                return HttpResponseRedirect(reverse('index'))
            else:
                return HttpResponseRedirect("ACCOUNT NOT ACTIVE")
        else:
            print("Someone tried to login and failed")
            print("Username: {} and password {}".format(username,password))
            return HttpResponse("invalid login details supplied")

    else:
        return render(request, 'ac_app/login.html',{})

def import_excel(request):

    #add validators if you want
    if  "GET" == request.method:
        return render(request, 'ac_app/import_excel.html', {})

    if request.method == 'POST':  
        excel_file = request.FILES["excel_file"]
        tolerance = 5
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet3 = workbook['Targets']
        sheet1 = workbook['Lease information']
        planned_wells = {}
        i = 0
        tot_wells = 0
        for row in range(65, sheet3.max_row +1):
            if sheet3['A' + str(row)].value == 1:
                tot_wells += 1

        for row in range(65, sheet3.max_row +1):
            if sheet3['A' + str(row)].value == 1:
                str_name = sheet3['B'+ str(row)].value
                y_shl = round(sheet3['F'+ str(row)].value,4)
                x_shl = round(sheet3['E'+ str(row)].value,4)
                kop_y = round(sheet3['F'+ str(row +1)].value,4)
                kop_x = round(sheet3['E'+ str(row +1)].value,4)
                az = sheet3['T'+ str(row)].value
                Well_name = str_name.split('#')[0][3:]+ str_name.split('#')[1][1]
                planned_wells[Well_name] = dict()
                if sheet3['D' + str(row + 2)].value == 'Heel':
                    heel_x = round(sheet3['E' + str(row + 2)].value,4)
                    heel_y = round(sheet3['F' + str(row + 2)].value,4)
                if sheet3['D' + str(row + 3)].value == 'Toe':
                    toe_x = round(sheet3['E' + str(row + 3)].value,4)
                    toe_y = round(sheet3['F' + str(row + 3)].value,4)
                #build analysis
                # ---------------------------------------------------------------------------------------------------------------------------
                #note (remember for simplistic reasons az on sheet is not true az relative to grid north or mag north it simply is a geometric degree)
                #ns build conditions
                # ---------------------------------------------------------------------------------------------------------------------------
                if ( 90-tolerance <= az  and  az <= 90+tolerance )  or ( -90-tolerance <= az and az <= -90+tolerance):
                    direction = 'N-S'
                    if az <= (-88) and y_shl > kop_y:
                        build = 'frontbuild north to south'
                    elif az >= (88) and kop_y > y_shl:
                        build = 'frontbuild south to north'
                    else:
                        build = ''
                #ew build conditions
                # ---------------------------------------------------------------------------------------------------------------------------
                elif (0-tolerance <= az and az <= 3+tolerance) or (180-tolerance <= az and az <= 180+tolerance) or (-180-tolerance <= az and az <= -180+tolerance):
                    direction = 'E-W'
                    if  (0-tolerance <= az and az <= 3+tolerance) and x_shl < kop_x:
                        build = 'frontbuild west to east'
                    elif (180-tolerance <= az and az <= 180+tolerance) and x_shl > kop_x:
                        build = 'frontbuild east to west'
                    else:
                        build = ''

                planned_wells[Well_name] = [heel_x, heel_y, toe_x, toe_y, y_shl, kop_y, build, x_shl, kop_x, direction]
                #planned_wells[Well_name]['heel_x'] = heel_x
                #planned_wells[Well_name]['heel_y'] = heel_y
                #planned_wells[Well_name]['toe_x'] = toe_x
                #planned_wells[Well_name]['toe_y'] = toe_y
                #planned_wells[Well_name]['y_shl'] = y_shl
                #planned_wells[Well_name]['kop_y'] = kop_y

                #planned_wells[Well_name]['build'] = build
                #planned_wells[Well_name]['x_shl'] = x_shl
                #planned_wells[Well_name]['kop_x'] = kop_x
                #planned_wells[Well_name]['direction'] = direction

                i += 1

                if i > tot_wells:
                    break
                # ---------------------------------------------------------------------------------------------------------------------------
        elevation = sheet1['H14'].value
        formation_tops = {}
        for row in range(27, sheet1.max_row +1):
            bench_tops = sheet1['G'+ str(row)].value
            tvd_tops = sheet1['I' + str(row)].value
            formation_tops[bench_tops] = tvd_tops
            if (sheet1['G'+ str(row + 1)].value) == '' or (sheet1['G'+ str(row + 1)].value) == None:
                break
        for k,v in formation_tops.items():
            if v == elevation:
                del formation_tops[k]

        request.session['planned_wells'] = planned_wells
        request.session['formation_tops'] = formation_tops
        #context = {"message": "Planning Sheet has finished Uploading Please upload the Well Inventory.."}

    return render(request, 'ac_app/import_excel.html',{'planned_wells':planned_wells}, {'formation_tops':formation_tops})
    

def import_excel2(request):

    #add validators if you want
    if  "GET" == request.method:
        return render(request, 'ac_app/import_excel.html', {})

    if request.method == 'POST':  
        excel_file = request.FILES["excel_file2"]
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet2 = workbook['Compiled list']
        worksheet = workbook['Data']
        #General array ---------------------------------------------------------------------------------------------------------------------------------
        dont_want_arr = ['', None]
        #---------------------------------------------------------------------------------------------------------------------------
        #functions for sf calcs
        # --------------------------------------------------------------------------------------------------------------------------
        def sp_nums(tvd):
            x = list(uncert.keys())
            sp_nums = []
            sp_hash = {}
            if tvd in x:
                return tvd
            else:
                x.append(tvd)
                x.sort()
                for idx, el in enumerate(x):
                    if el == tvd and tvd != x[-1]:
                        sp_nums.append(x[idx - 1])
                        sp_nums.append(x[idx + 1])
                    elif el == tvd:
                        a = x[idx - 2]
                        b = x[idx - 1]
                        sp_hash[tvd] = [a, b]
                        return sp_hash
                return sp_nums

        #lateral wells calc

        def sp_numslat(lat_dist):
            x = list(lat_uncert.keys())
            sp_nums = []
            sp_hash = {}
            if lat_dist in x:
                return lat_uncert[lat_dist][2]
            else:
                x.append(lat_dist)
                x.sort()
                for idx, el in enumerate(x):
                    if el == lat_dist and lat_dist != x[-1]:
                        sp_nums.append(x[idx - 1])
                        sp_nums.append(x[idx + 1])
                    elif el == lat_dist:
                        a = x[idx - 2]
                        b = x[idx - 1]
                        sp_hash[lat_dist] = [a, b]
                        return sp_hash[lat_dist]
                return sp_nums
        # this is for only wells with  str must be = no_inc or inc


        def vert_interp(tvd, str1):
            if str1 == 'INC':
                x = tvd
                data = sp_nums(tvd)
                if type(data) == list:
                    x1 = data[1]
                    x0 = data[0]
                    y1 = uncert[x1]
                    y0 = uncert[x0]
                    y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                    y = round(y,4)
                    return y
                elif type(data) == dict:
                    x1 = data[tvd][1]
                    x0 = data[tvd][0]
                    y1 = uncert[x1]
                    y0 = uncert[x0]
                    y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                    y = round(y,4)
                    return y
                else:
                    return uncert[tvd]
            else:
                x = tvd
                data = sp_nums(tvd)
                if type(data) == list:
                    x1 = data[1]
                    x0 = data[0]
                    y1 = uncert1[x1]
                    y0 = uncert1[x0]
                    y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                    y = round(y,4)
                    return y
                elif type(data) == dict:
                    x1 = data[tvd][1]
                    x0 = data[tvd][0]
                    y1 = uncert1[x1]
                    y0 = uncert1[x0]
                    y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                    y = round(y,4)
                    return y
                else:
                    return uncert1[x]


        def lat_interp(lat_dist, direction, met):
            #case 1 example 1 of PRNM ac guidelines
            #case 2 example 2 of PRNM ac guidelines
            if met == 'case 1':
                if direction == 'N-S':
                    x = lat_dist
                    data = sp_numslat(lat_dist)
                    if type(data) == list:
                        x1 = data[1]
                        x0 = data[0]
                        y1 = lat_uncert[x1][1]
                        y0 = lat_uncert[x0][1]
                        y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                        y = round(y,4)
                        return y
                    elif type(data) == dict:
                            x1 = data[lat_dist][1]
                            x0 = data[lat_dist][0]
                            y1 = lat_uncert[x1][1]
                            y0 = lat_uncert[x0][1]
                            y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                            y = round(y,4)
                            return y
                    else:
                        return lat_uncert[lat_dist][1]
                #accounting for E_W wells
                else:
                    x = lat_dist
                    data = sp_numslat(lat_dist)
                    if type(data) == list:
                        x1 = data[1]
                        x0 = data[0]
                        y1 = lat_uncert[x1][0]
                        y0 = lat_uncert[x0][0]
                        y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                        y = round(y,4)
                        return y
                    elif type(data) == dict:
                            x1 = data[lat_dist][1]
                            x0 = data[lat_dist][0]
                            y1 = lat_uncert[x1][0]
                            y0 = lat_uncert[x0][0]
                            y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                            y = round(y,4)
                            return y
                    else:
                        return lat_uncert[lat_dist][0]

            #note case 2 2 lateral wells inputs(2 lat dist) ouptput is 2 (tvd uncert)
            else:
                x = lat_dist
                data = sp_numslat(lat_dist)
                if type(data) == list:
                    x1 = data[1]
                    x0 = data[0]
                    y1 = lat_uncert[x1][2]
                    y0 = lat_uncert[x0][2]
                    y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                    y = round(y,4)
                    return y
                elif type(data) == dict:
                        x1 = data[lat_dist][1]
                        x0 = data[lat_dist][0]
                        y1 = lat_uncert[x1][2]
                        y0 = lat_uncert[x0][2]
                        y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                        y = round(y,4)
                        return y
                else:
                    return lat_uncert[lat_dist][2]


        #selects only rows containing key in specific cell
        def selectdict(dict2, idx, key):
            new_dict = dict((k, v) for k, v in dict2.iteritems() if v[idx] == key)
            return new_dict

        #for making a dictionary excluding all apis which satisfy a specific condition for instance key = 'HORIZONTAL' new dict would have every api that isnt a horizontal


        def unselectdict(dict2, idx, key):
            new_dict = dict((k, v) for k, v in dict2.iteritems() if v[idx] != key)
            return new_dict

        # if key array = [0,none, ''] and various rows have this in a specific cell value for a specific row this api will be ignored


        def multiunselectdict(dict2, idx, keyarr):
            new_dict = dict((k, v) for k, v in dict2.iteritems() if v[idx] not in keyarr)
            return new_dict

        #--------------------------------------------------------------------------------------------------------------
        #parameters for Arcpy functions
        zonesG = r'\\ohoagis\gis\SurfacePlanning\ProjectData\_Enterprise\PYTHON_PING\StatePlaneZones.gdb\StatePlaneZones_WKID'
        zone_lyrG = MakeFeatureLayer_management(zonesG)

        hpdi_wells = r'\\ohoagis\gis\SurfacePlanning\SDE_GDB_Connect\HOSDEP3_sde_user@DEFAULT_oxydom.sde\HPDI.DI_WELLS_DIRECTIONALS'

        quearybase = "API ='"
        #------------------------------------------------------------------------------------------------------------------
        def getWKID2(SHAPE):
            """
                Function is used to calculate the CRS code for a shape@
            """
            zone_selection = SelectLayerByLocation_management(zone_lyrG, 'INTERSECT',
                                                            SHAPE, selection_type='NEW_SELECTION')
            WKID = [row[0] for row in SearchCursor(zone_selection, ['WKID'])]
            return WKID[0]


        def main(APIList, SRCODE=None):
            """
                Function loops through list of APIs to find XY values for the features
                in the HPDI.DI_WELLS_DIRECTIONALS layer.

                Returns dictionary Dictionary[API]=[x.xxx,y.yyyy]

                For faster runt time, use SRCODE for reional searches where the SRCODE
                is known and constant


                --Important--
                    Function only finds first API that matches query.  If there are
                    multiple wells with the same API, only the first one will be returned

            """
            ApiDic = {}

            #Loop though API
            for api in APIList:
                #Build query string
                query = quearybase+str(api)+"'"

                #Import layer from SDE
                with SearchCursor(hpdi_wells, ["API", "SHAPE@"], query) as cursor:

                    #Loop though each feature found with query
                    for row in cursor:
                        Shape_points = []
                        
                        GeoShape = row[-1]

                        #Find CRS code from shape location and project
                        if not SRCODE:  # Was SRCODE provied?
                            SRCODE = getWKID2(GeoShape)
                        spat_ref = SpatialReference(SRCODE)
                        Pt_Proj = GeoShape.projectAs(spat_ref)

                        #Break geometry into points
                        for part in Pt_Proj:
                            for pnt in part:
                                Shape_points.append([pnt.X, pnt.Y])

                        #Save results to dictionary
                        ApiDic[api] = Shape_points
                        break

                    else:
                        #If query had no results, return none in dictionary
                        ApiDic[api] = None

            return ApiDic


        #-------------------------------------------------------------------------------------------------------------------------------------------------------------------
        #import Analagous well Uncertainty Data and Offset Well Data from Well Inventory

        uncert = OrderedDict()
        uncert1 = OrderedDict()
        i = 0
        #note uncert and uncert1 are actually lat-uncert, for vertical offset wells with no survey and a inc survey
        for row in range(5, worksheet.max_row +1):
            tvd = worksheet['A'+ str(row)].value
            no_survey = worksheet['B'+ str(row)].value
            inc_only = worksheet['C' + str(row)].value
            #gyro = worksheet['D' + str(row)].value
            uncert1[tvd]= no_survey
            uncert[tvd] = inc_only
            if worksheet['A'+ str(row +1)].value in dont_want_arr:
                break 

        lat_uncert = OrderedDict()
        i = 0
        #lat uncert are tvd uncertainties for lat wells (lat_dist)
        #*********** these uncertainties need to be refactored in all the code
        for row in range(5, worksheet.max_row +1):
            lat_length = worksheet['F'+ str(row)].value
            tvd_uncert = worksheet['I'+ str(row)].value
            ns_lat_uncert = worksheet['H' + str(row)].value
            ew_lat_uncert = worksheet['G' + str(row)].value
            lat_uncert[lat_length] = [ew_lat_uncert,ns_lat_uncert, tvd_uncert]
            if worksheet['F'+ str(row +1)].value in dont_want_arr:
                break

        for col in range(1, sheet2.max_column + 1):
            if (sheet2.cell(row=1, column=col).value == '' or sheet2.cell(row=1, column=col).value == None) and (sheet2.cell(row=1, column=col).value == None or sheet2.cell(row=1, column=col).value == ''):
                continue 
            elif 'API #' in str(sheet2.cell(row=1, column=col).value).strip().upper():
                api_col = get_column_letter(col)
            elif 'GIS X' in str(sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
                x_col = get_column_letter(col)
            elif 'GIS Y' in str(sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
                y_col = get_column_letter(col)
            elif 'STATUS' in str(sheet2.cell(row=1, column=col).value).strip().upper() and 'GIS' not in str(sheet2.cell(row=1, column=col).value).strip().upper():
                stat_col = get_column_letter(col)
            elif 'TVD (FT)' in str(sheet2.cell(row=1, column=col).value).strip().upper() and 'REFERENCE' not in str(sheet2.cell(row=1, column=col).value).strip().upper():
                tvd_col = get_column_letter(col)
            elif 'PROFILE' in str(sheet2.cell(row=1, column=col).value).strip().upper():
                pr_col = get_column_letter(col)
            elif 'SURVEY TYPE' in str(sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
                surv_col = get_column_letter(col)

        def offsetwells():
            loc_arr = ['PERM-LOC','AB-LOC']
            offset_wells_pre = OrderedDict()
            for row in range(2, sheet2.max_row +1):
                if (sheet2[stat_col + str(row)].value.upper() not in loc_arr) and (sheet2[stat_col + str(row)].value not in dont_want_arr):
                    api = sheet2[api_col + str(row)].value 
                    x = sheet2[x_col + str(row)].value
                    y = sheet2[y_col + str(row)].value
                    status = sheet2[stat_col + str(row)].value.upper()
                    tvd_offset = sheet2[tvd_col + str(row)].value
                    if sheet2[surv_col + str(row)].value in dont_want_arr:
                        continue 
                    surv_type = sheet2[surv_col + str(row)].value.upper()
                    profile = sheet2[pr_col + str(row)].value.upper()
                    offset_wells_pre[api] = [x,y,status,tvd_offset, surv_type,profile] 
                if sheet2['H'+ str(row +1)].value in dont_want_arr:
                    break
            return offset_wells_pre

        offset_wells_pre = offsetwells() 
        planned_wells = request.session.get('planned_wells', None)
        formation_tops = request.session.get('formation_tops', None)

        
        #print formation_tops
        form_list = sorted(list(formation_tops.values()))
        tvd_deep_for_inc_only = form_list[0] - 200
        #print tvd_deep_for_inc_only
        tvd_deep_for_no_surv = form_list[0] - 1000
        #print tvd_deep_for_no_surv

        offset_wells_pre_2 = OrderedDict()
        for k, v in offset_wells_pre.items():
            if (v[3] >= tvd_deep_for_inc_only and ('INC ONLY' in v[4])) or (v[3] >= tvd_deep_for_no_surv and ('NO' in v[4])):
                offset_wells_pre_2[k] = v
        

        offset_laterals_pre = selectdict(offset_wells_pre_2,idx =5 ,key = 'HORIZONTAL' )
        offset_lat = list(offset_laterals_pre.keys())
        offset_directional = selectdict(offset_wells_pre_2,idx =5 ,key = 'DIRECTIONAL' )

        offset_wells_hash = offset_wells_pre_2

        MyXY=main(offset_lat, 32012) #2
        #print MyXY

        laterals_with_no_eot = {}
        offset_laterals = OrderedDict()
        for k,v in offset_laterals_pre.iteritems():
            for k1,v1 in MyXY.iteritems():
                if (v1 != None) and (k1 == k):
                    diff_x = abs(v[0] - v1[0][0])
                    diff_y = abs(v[1] - v1[0][1])
                    diff_x1 = abs(v[0] - v1[1][0])
                    diff_y1 = abs(v[1] - v1[1][1])
                    r = math.sqrt( diff_x**2 + diff_y**2)
                    r1 = math.sqrt(diff_x1**2 + diff_y1**2)
                    if r <= 200 and r1 >= 1000:
                        xf = v1[1][0]
                        yf = v1[1][1]
                        theta = float(math.atan( float( (diff_y1)/(diff_x1))) * float(180/ (float(math.pi))))
                        theta = round(theta,4)
                        tolerance = 15
                        if ( 90-tolerance <= theta  and  theta <= 90+tolerance )  or ( -90-tolerance <= theta and theta <= -90+tolerance):
                            direction = 'N-S'
                            offset_laterals[k] = [v[0],v[1],v[2],v[3],v[4],v[5], direction, xf, yf]

                        elif ( 0-tolerance <= theta and theta <= 3+tolerance) or (180-tolerance <= theta and theta <= 180+tolerance):
                            direction = 'E-W'
                            offset_laterals[k] = [v[0],v[1],v[2],v[3],v[4],v[5], direction, xf, yf]
                elif (v1 == None) and (k1 == k):
                    laterals_with_no_eot[k] = [v[0],v[1],v[2],v[3],v[4],v[5]]

        # Well Calculations Case 1 ( Deep Well and  planned lateral Well)
        # ------------------------------------------------------------------ -------------------------------------------------------------------------------------------------
        api = list(offset_wells_hash.keys()) #3
        wellnames = list(planned_wells.keys())
        wells = []
        select_api = []
        surv = []
        select_x = []
        select_y = []
        select_tvd = []
        sf_one = []
        sf_two = []
        ctr_to_ctr = []
        stat = []
        gyro = []
        pr = []
        lat = []
        wells_1 = []
        select_api_1 = []
        surv_1 = []
        select_x_1 = []
        select_y_1 = []
        select_tvd_1 = []
        sf_one_1 = []
        sf_two_1 = []
        ctr_to_ctr_1 = []
        stat_1 = []
        gyro = []
        pr_1 = []
        lat_1 = []
        case = []
        case_1 = []
        #print formation_tops
        #print planned_wells
        for el1 in wellnames:
            toe_x = planned_wells[el1][2]
            toe_y = planned_wells[el1][3]
            heel_y = planned_wells[el1][1]
            heel_x = planned_wells[el1][0]
            y_shl = planned_wells[el1][4]
            kop_y = planned_wells[el1][5]
            build = planned_wells[el1][6]
            x_shl = planned_wells[el1][7]
            kop_x = planned_wells[el1][8]
            direction_planned_well = planned_wells[el1][9]
        # N_S Well Calculations-------------------------------------------------------------------------------------------------------------------------------------------------------------------
            if direction_planned_well == 'N-S':
                for k,v in formation_tops.items():
                    if k in el1: 
                        planned_well_tvd = v
                        break
                for el in api: #4
                    status = offset_wells_hash[el][2]
                    survey = offset_wells_hash[el][4]
                    x_offset = offset_wells_hash[el][0]
                    y_offset = offset_wells_hash[el][1]
                    tvd_offset = offset_wells_hash[el][3]
                    profile = offset_wells_hash[el][5]
                    if build == '':
                        if (kop_y < y_offset and y_offset < heel_y) or (heel_y < y_offset and y_offset < kop_y):
                            lat_dist_fr_heel = 0 
                        else: 
                            lat_dist_fr_heel = abs(float(heel_y - y_offset))
                            lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                        lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                        if 'INC ONLY' in survey :
                            str1 = 'INC'
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( float((vert_uncert)**(2)) + float((lat_uncertainty)**(2)) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #5
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #6
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #7
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #8
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                        
                            elif tvd_offset <= planned_well_tvd:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ((  (toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #9
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #10
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #11
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #12
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                
                            
                        else:
                            str1 = ''
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf, 4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf, 4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #13
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #14
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #15
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #16
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})


                            elif  planned_well_tvd >= tvd_offset:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ( ((toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #17
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #18
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                            
                                else:
                                    ctr_to_ctr_dist = float( ( ((heel_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #19
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #20
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                    

                    
                    else:          
                        if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                            lat_dist_fr_heel = 0
                        else :
                            lat_dist_fr_heel = abs(float(heel_y - y_offset))
                            lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                        lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                        if  'INC ONLY' in survey:
                            str1 = 'INC'
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf, 4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #21
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #22
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist, 4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #23
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #24
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                        
                            elif tvd_offset <= planned_well_tvd:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ((  (toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist, 4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #25
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #26
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist, 4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #27
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #28
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                
                            
                        else :
                            str1 = ''
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf, 4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #29
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #30
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #31
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #32
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})


                            elif  planned_well_tvd >= tvd_offset:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ( ((toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #33
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #34
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                            
                                else:
                                    ctr_to_ctr_dist = float( ( ((heel_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el) #35
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el) #36
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
        # E_W Calculations ------------------------------------------------------------------------------------------------------------------------------------------------------------------- 
            elif direction_planned_well == 'E-W':
                for k,v in formation_tops.items():
                    if k in el1: 
                        planned_well_tvd = v
                        break
                for el in api: 
                    status = offset_wells_hash[el][2]
                    survey = offset_wells_hash[el][4]
                    x_offset = offset_wells_hash[el][0]
                    y_offset = offset_wells_hash[el][1]
                    tvd_offset = offset_wells_hash[el][3]
                    profile = offset_wells_hash[el][5]
                    if build == '':
                        if (kop_x < x_offset and x_offset < heel_x) or (heel_x < x_offset and x_offset < kop_x):
                            lat_dist_fr_heel = 0 
                        else: 
                            lat_dist_fr_heel = abs(float(heel_x - x_offset))
                            lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                        lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                        if  'INC ONLY' in survey:
                            str1 = 'INC'
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf, 4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf, 4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                        
                            elif tvd_offset <= planned_well_tvd:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ((  (toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                
                            
                        else:
                            str1 = ''
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd, str1)
                                one_sf = float(
                                    (((vert_uncert)**(2) + (lat_uncertainty)**(2))**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})


                            elif  planned_well_tvd >= tvd_offset:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ( ((toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                            
                                else:
                                    ctr_to_ctr_dist = float( ( ((heel_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                    

                    
                    else:          
                        if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                            lat_dist_fr_heel = 0
                        else :
                            lat_dist_fr_heel = abs(float(heel_y - y_offset))
                            lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                        lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                        if  'INC ONLY' in survey:
                            str1 = 'INC'
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf, 4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf, 4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                        
                            elif tvd_offset <= planned_well_tvd:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ((  (toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                
                            
                        else :
                            str1 = ''
                            vert_uncert = vert_interp(tvd_offset,str1)  
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            one_sf = round(one_sf,4)
                            two_sf = float(one_sf*2)
                            two_sf = round(two_sf,4)
                            if tvd_offset >= planned_well_tvd: 
                                vert_uncert = vert_interp(planned_well_tvd,str1)  
                                one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                                else:
                                    ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})


                            elif  planned_well_tvd >= tvd_offset:
                                if lat_dist_fr_heel >= 5000:
                                    ctr_to_ctr_dist = float( ( ((toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                            
                                else:
                                    ctr_to_ctr_dist = float( ( ((heel_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                    ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                    if two_sf >=  ctr_to_ctr_dist:
                                        need_gyro = 'True'
                                        wells.append(el1)
                                        select_api.append(el)
                                        surv.append(survey)
                                        stat.append(status)
                                        select_x.append(x_offset)
                                        select_y.append(y_offset)
                                        select_tvd.append(tvd_offset)
                                        sf_one.append(one_sf)
                                        sf_two.append(two_sf)
                                        ctr_to_ctr.append(ctr_to_ctr_dist)
                                        pr.append(profile)
                                        lat.append(lat_dist_fr_heel)
                                        case.append(1)
                                        anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'lateral_distance': lat,'case':case})
                                    else:
                                        wells_1.append(el1)
                                        select_api_1.append(el)
                                        surv_1.append(survey)
                                        stat_1.append(status)
                                        select_x_1.append(x_offset)
                                        select_y_1.append(y_offset)
                                        select_tvd_1.append(tvd_offset)
                                        sf_one_1.append(one_sf)
                                        sf_two_1.append(two_sf)
                                        ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                        pr_1.append(profile)
                                        lat_1.append(lat_dist_fr_heel)
                                        case_1.append(1)
                                        anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
    
    
                    
    # Well Calculations Case 2 ( 2 orthogonal laterals)
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
    api = offset_laterals.keys()
    wellnames = planned_wells.keys()
    for el1 in wellnames:
        toe_x = planned_wells[el1][2]
        toe_y = planned_wells[el1][3]
        heel_y = planned_wells[el1][1]
        heel_x = planned_wells[el1][0]
        y_shl = planned_wells[el1][4]
        kop_y = planned_wells[el1][5]
        build = planned_wells[el1][6]
        x_shl = planned_wells[el1][7]
        kop_x = planned_wells[el1][8]
        direction_planned_well = planned_wells[el1][9]
        #NS Planned well Calculations ------------------------------------------------------------------------------------------------------------------------------------------------------------------- 
        if direction_planned_well == 'N-S':
            for k,v in formation_tops.items():
                if k in el1: 
                    planned_well_tvd = v
                    break
            for el in api:
                status = offset_laterals[el][2]
                survey = offset_laterals[el][4]
                x_offset = offset_laterals[el][0]
                y_offset = offset_laterals[el][1]
                tvd_offset = offset_laterals[el][3]
                profile = offset_laterals[el][5]
                direction = offset_laterals[el][6]
                x_f = offset_laterals[el][7]
                y_f = offset_laterals[el][8]
                if (x_offset < heel_x and heel_x < x_f) or ( x_f < heel_x and heel_x < x_offset):
                    if direction == 'E-W':
                        if build == '':
                            if (kop_y < y_offset and  y_offset < toe_y) or (toe_y < y_offset and y_offset < kop_y): # basically saying if your within the y domain of the planned well
                                if (kop_y < y_offset and y_offset < heel_y) or (heel_y < y_offset and y_offset < kop_y):
                                    lat_dist_fr_heel = 0 
                                else: 
                                    lat_dist_fr_heel = abs(float(y_offset - heel_y))
                                    lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                                lat_dist_intersect = abs(float(heel_x - x_offset))
                                lat_dist_intersect = round(lat_dist_intersect,4)
                                lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'lateral_distance': lat,'case':case}) 
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                        else:
                            if (y_shl < y_offset and y_offset < toe_y) or (toe_y < y_offset and y_offset < y_shl): # basically saying if you meet front build case and your not in its ydomain  dont do any calcs
                                if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                                    lat_dist_fr_heel = 0
                                else :
                                    lat_dist_fr_heel = abs(float(heel_y - y_offset))
                                    lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                                lat_dist_intersect = abs(float(heel_x - x_offset))
                                lat_dist_intersect = round(lat_dist_intersect,4)
                                lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'lateral_distance': lat,'case':case}) 
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
        # E_W Planned well Calculations ------------------------------------------------------------------------------------------------------------------------------------------------------------------- 
        elif direction_planned_well == 'E-W':
            for k,v in formation_tops.items():
                if k in el1: 
                    planned_well_tvd = v
                    break
            for el in api:
                status = offset_laterals[el][2]
                survey = offset_laterals[el][4]
                x_offset = offset_laterals[el][0]
                y_offset = offset_laterals[el][1]
                tvd_offset = offset_laterals[el][3]
                profile = offset_laterals[el][5]
                direction = offset_laterals[el][6]
                x_f = offset_laterals[el][7]
                y_f = offset_laterals[el][8]
                if (y_offset < heel_y and heel_y < y_f) or ( y_f < heel_y and heel_y < y_offset):
                    if direction == 'N-S':
                        if build == '':
                            if (kop_x < x_offset and  x_offset < toe_x) or (toe_x < x_offset and x_offset < kop_x): # basically saying if your within the y domain of the planned well
                                if (kop_x < x_offset and x_offset < heel_x) or (heel_x < x_offset and x_offset < kop_x):
                                    lat_dist_fr_heel = 0 
                                else: 
                                    lat_dist_fr_heel = abs(float(x_offset - heel_x))
                                    lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                                lat_dist_intersect = abs(float(heel_y - y_offset))
                                lat_dist_intersect = round(lat_dist_intersect,4)
                                lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'lateral_distance': lat,'case':case}) 
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                        else:
                            if (x_shl < x_offset and x_offset < toe_x) or (toe_x < x_offset and x_offset < x_shl): # basically saying if you meet front build case and are within planned wells domain
                                if (heel_x < x_offset and x_offset < x_shl) or (x_shl < x_offset and x_offset < heel_x): # if intersect is between shl and heel lat dist from heel is 0 meaning still in build
                                    lat_dist_fr_heel = 0
                                else :
                                    lat_dist_fr_heel = abs(float(heel_x - x_offset))
                                    lat_dist_fr_heel = round(lat_dist_fr_heel,4)
                                lat_dist_intersect = abs(float(heel_y - y_offset))
                                lat_dist_intersect = round(lat_dist_intersect,4)
                                lat_uncertainty = lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                one_sf = round(one_sf,4)
                                two_sf = float(one_sf*2)
                                two_sf = round(two_sf,4)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                ctr_to_ctr_dist = round(ctr_to_ctr_dist,4)
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, 'sf_one': sf_one, 'sf_two': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'lateral_distance': lat,'case':case}) 
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, 'sf_one': sf_one_1, 'sf_two': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'lateral_distance': lat_1,'case':case_1})
                        
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Labels
    #--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    #create labels
        
        
    
    
    ac_1 = anti_collision_sf
    

    #ac_2 = anti_collision_sf_2.to_dict()        
    #return render(request, 'ac_app/import_excel.html',ac_1, ac_2)
    
        
    ac_1.rename(columns={'Well':'well', 'API':'api','Survey':'survey','X':'x','Y':'y','TVD':'tvd','Center to Center Distance':'center_center_distance','Profile':'traj'}, 
                inplace=True)
    ac_new = ac_1
    request.session['ac_new'] = ac_new
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    path = os.path.join(BASE_DIR, 'static')
    os.chdir(path)
    ac_1.to_csv('ac_1.csv',index=False)
    
        
    fj = pyexcel.save_as(file_name ='ac_1.csv', dest_file_name='ac_1.xlsx')
    #file_h = pyexcel.save_book_as(array=ac_1, dest_file_name=r"C:\Users\cubasg\Desktop\My_Django_Stuff\practice\ac_tool\static\ac_1.csv")
    #file_h = excel.make_response(pyexcel.get_book(file_name='ac_1.csv'), "csv", file_name="ac_1") #this is used to download a file this exact one **
    file_h = pyexcel.get_sheet(file_name='ac_1.xlsx', name_columns_by_row=0)
    file_h.save_to_django_model(
         model=AcFlaggedWell, 
         initializer=None, 
         #mapdict=['sf_one', 'sf_two', 'api','center_center_distance','traj','survey','tvd','well','x','y','case','lateral_distance'])
         mapdict=['api','center_center_distance','traj','survey','tvd','well','x','y','case','lateral_distance','sf_one','sf_two'])
    
    return render(request,'ac_app/custom-handson-table.html',  {'ac_new': ac_new})

def download(request):

    ac_1 = request.session.get('ac_new', None)
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    path = os.path.join(BASE_DIR, 'static')
    os.chdir(path)

    ac_1.to_csv('ac_new.csv')
    pyexcel.save_as(file_name='ac_new.csv', dest_file_name='ac_new.xlsx')
    excel.make_response(pyexcel.get_book(file_name='ac_new.xlsx'), "xlsx", file_name="ac_new")

def handson_table(request):
    return excel.make_response_from_a_table(AcFlaggedWell, 'handsontable.html')
    

def embed_handson_table_from_a_single_table(request):
    """Renders two table in a handsontable"""
    content = excel.pe.save_as(
        model=AcFlaggedWell,
        dest_file_type='handsontable.html',
        dest_embed=True)
    content.seek(0)
    return render(request,'ac_app/custom-handson-table.html',{'handsontable_content': content.read()})
