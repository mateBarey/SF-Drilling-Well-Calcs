#-------------------------------------------------------------------------------
# Name:        Sfcalcnew
# Purpose:
#
# Author:      George Raul Cubas
#
# Created:     X/XX/XXXX
# Copyright:   (c) George Raul Cubas
# Licence:     <MIT>
#-------------------------------------------------------------------------------
# Program uses ISCWSA Error Model for calculating a separation factor for Laterals and Vertical Wells
#**************************************************************************************
# System Setup
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import  get_column_letter
import pandas as pd
import os
from  Sfimportnew import Plannedwells
import math
import numpy as np
from collections import OrderedDict
import time
#**************************************************************************************

class Sfcalcnew():
  def __init__(self,wellsfile,plannedwellfile,outputfilename):
    self.wellsfile = wellsfile
    self.plannedwellfile = plannedwellfile
    self.outputfilename = outputfilename
    self.workbook = openpyxl.load_workbook(self.wellsfile, data_only=True)
    self.sheet2 = self.workbook['Compiled list']
    self.worksheet = self.workbook['Data']
    self.dont_want_arr = ['',None]
  def uncertainty(self):

      uncert = OrderedDict()
      uncert1 = OrderedDict()
      i = 0
      for row in range(5, self.worksheet.max_row +1):
        tvd = self.worksheet['A'+ str(row)].value
        no_survey = self.worksheet['B'+ str(row)].value
        inc_only = self.worksheet['C' + str(row)].value
        uncert1[tvd] = no_survey
        uncert[tvd] = inc_only
        if self.worksheet['A'+ str(row +1)].value in self.dont_want_arr:
            break

      lat_uncert = OrderedDict()
      i = 0
      for row in range(5, self.worksheet.max_row +1):
        lat_length = self.worksheet['F'+ str(row)].value
        tvd_uncert = self.worksheet['I'+ str(row)].value
        ns_lat_uncert = self.worksheet['H' + str(row)].value
        ew_lat_uncert = self.worksheet['G' + str(row)].value
        lat_uncert[lat_length] = [ew_lat_uncert,ns_lat_uncert, tvd_uncert]
        if self.worksheet['F'+ str(row +1)].value in self.dont_want_arr:
            break
      return uncert, uncert1, lat_uncert

  def offsetwells(self):
      loc_arr = ['PERM-LOC','AB-LOC']
      offset_wells_pre = OrderedDict()
      for col in range(1, self.sheet2.max_column + 1):
          if (self.sheet2.cell(row=1, column=col).value == '' or self.sheet2.cell(row=1, column=col).value == None) and (self.sheet2.cell(row=1, column=col).value == None or self.sheet2.cell(row=1, column=col).value == ''):
              continue
          elif 'API #' in str(self.sheet2.cell(row=1, column=col).value).strip().upper():
              api_col = get_column_letter(col)
          elif 'GIS X' in str(self.sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
              x_col = get_column_letter(col)
          elif 'GIS Y' in str(self.sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
              y_col = get_column_letter(col)
          elif 'STATUS' in str(self.sheet2.cell(row=1, column=col).value).strip().upper() and 'GIS' not in str(self.sheet2.cell(row=1, column=col).value).strip().upper():
              stat_col = get_column_letter(col)
          elif 'TVD (FT)' in str(self.sheet2.cell(row=1, column=col).value).strip().upper() and 'REFERENCE' not in str(self.sheet2.cell(row=1, column=col).value).strip().upper():
              tvd_col = get_column_letter(col)
          elif 'PROFILE' in str(self.sheet2.cell(row=1, column=col).value).strip().upper():
              pr_col = get_column_letter(col)
          elif 'SURVEY TYPE' in str(self.sheet2.cell(row=1, column=col).value).strip().upper().replace('\n', ''):
              surv_col = get_column_letter(col)
      for row in range(2, self.sheet2.max_row +1):
          if (self.sheet2[stat_col + str(row)].value.upper() not in loc_arr) and (self.sheet2[stat_col + str(row)].value not in self.dont_want_arr):
              api = self.sheet2[api_col + str(row)].value
              x = self.sheet2[x_col + str(row)].value
              y = self.sheet2[y_col + str(row)].value
              status = self.sheet2[stat_col + str(row)].value.upper()
              tvd_offset = self.sheet2[tvd_col + str(row)].value
              if self.sheet2[surv_col + str(row)].value in self.dont_want_arr:
                  continue
              surv_type = self.sheet2[surv_col + str(row)].value.upper()
              profile = self.sheet2[pr_col + str(row)].value.upper()
              offset_wells_pre[api] = [x,y,status,tvd_offset, surv_type,profile] #1
          if self.sheet2['H'+ str(row +1)].value in self.dont_want_arr:
              break
      return offset_wells_pre

  # Useful helper methods
  def sp_nums(self,tvd):
      self.tvd = tvd
      uncert, uncert1, lat_uncert = self.uncertainty()
      x = list(uncert.keys())
      sp_nums = []
      sp_hash = {}
      if self.tvd in x:
          return self.tvd
      else:
          x.append(self.tvd)
          x.sort()
          for idx, el in enumerate(x):
              if el == self.tvd and self.tvd != x[-1]:
                  sp_nums.append(x[idx - 1])
                  sp_nums.append(x[idx + 1])
              elif el == tvd:
                  a = x[idx - 2]
                  b = x[idx - 1]
                  sp_hash[self.tvd] = [a, b]
                  return sp_hash
          return sp_nums

  #lateral wells calc

  def sp_numslat(self,lat_dist):
      self.lat_dist = lat_dist
      uncert, uncert1, lat_uncert = self.uncertainty()
      x = list(lat_uncert.keys())
      sp_nums = []
      sp_hash = {}
      if lat_dist in x:
          return lat_uncert[self.lat_dist][2]
      else:
          x.append(self.lat_dist)
          x.sort()
          for idx, el in enumerate(x):
              if el == self.lat_dist and self.lat_dist != x[-1]:
                  sp_nums.append(x[idx - 1])
                  sp_nums.append(x[idx + 1])
              elif el == self.lat_dist:
                  a = x[idx - 2]
                  b = x[idx - 1]
                  sp_hash[self.lat_dist] = [a, b]
                  return sp_hash[self.lat_dist]
          return sp_nums
  # this is for only wells with  str must be = no_inc or inc


  def vert_interp(self,tvd, str1):
      self.tvd = tvd
      self.str1 = str1
      uncert, uncert1, lat_uncert = self.uncertainty()
      if self.str1 == 'INC':
          x = self.tvd
          data = self.sp_nums(self.tvd)
          if type(data) == list:
              x1 = data[1]
              x0 = data[0]
              y1 = uncert[x1]
              y0 = uncert[x0]
              y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
              return y
          elif type(data) == dict:
              x1 = data[self.tvd][1]
              x0 = data[self.tvd][0]
              y1 = uncert[x1]
              y0 = uncert[x0]
              y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
              return y
          else:
              return uncert[tvd]
      else:
          x = self.tvd
          data = self.sp_nums(self.tvd)
          if type(data) == list:
              x1 = data[1]
              x0 = data[0]
              y1 = uncert1[x1]
              y0 = uncert1[x0]
              y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
              return y
          elif type(data) == dict:
              x1 = data[self.tvd][1]
              x0 = data[self.tvd][0]
              y1 = uncert1[x1]
              y0 = uncert1[x0]
              y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
              return y
          else:
              return uncert1[x]


  def lat_interp(self,lat_dist, direction, met):
      #case 1 example 1 of PRNM ac guidelines
      #case 2 example 2 of PRNM ac guidelines
      self.lat_dist = lat_dist
      self.direction = direction
      self.met = met
      uncert, uncert1, lat_uncert = self.uncertainty()
      if self.met == 'case 1':
          if self.direction == 'N-S':
              x = self.lat_dist
              data = self.sp_numslat(self.lat_dist)
              if type(data) == list:
                  x1 = data[1]
                  x0 = data[0]
                  y1 = lat_uncert[x1][1]
                  y0 = lat_uncert[x0][1]
                  y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                  return y
              elif type(data) == dict:
                      x1 = data[self.lat_dist][1]
                      x0 = data[self.lat_dist][0]
                      y1 = lat_uncert[x1][1]
                      y0 = lat_uncert[x0][1]
                      y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                      return y
              else:
                  return lat_uncert[self.lat_dist][1]
          #accounting for E_W wells
          else:
              x = self.lat_dist
              data = self.sp_numslat(self.lat_dist)
              if type(data) == list:
                  x1 = data[1]
                  x0 = data[0]
                  y1 = lat_uncert[x1][0]
                  y0 = lat_uncert[x0][0]
                  y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                  return y
              elif type(data) == dict:
                      x1 = data[self.lat_dist][1]
                      x0 = data[self.lat_dist][0]
                      y1 = lat_uncert[x1][0]
                      y0 = lat_uncert[x0][0]
                      y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                      return y
              else:
                  return lat_uncert[self.lat_dist][0]

      #note case 2 2 lateral wells inputs(2 lat dist) ouptput is 2 (tvd uncert)
      else:
          x = self.lat_dist
          data = self.sp_numslat(lat_dist)
          if type(data) == list:
              x1 = data[1]
              x0 = data[0]
              y1 = lat_uncert[x1][2]
              y0 = lat_uncert[x0][2]
              y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
              return y
          elif type(data) == dict:
                  x1 = data[self.lat_dist][1]
                  x0 = data[self.lat_dist][0]
                  y1 = lat_uncert[x1][2]
                  y0 = lat_uncert[x0][2]
                  y = float((float((y1-y0))/float((x1-x0))*(x-x0) + y0))
                  return y
          else:
              return lat_uncert[self.lat_dist][2]

  #selects only rows containing key in specific cell
  def selectdict(self,dict2, idx, key):
      self.dict2 = dict2
      self.idx = idx
      self.key = key
      new_dict = dict((k, v) for k, v in self.dict2.items() if v[self.idx] == self.key)
      return new_dict

  #for making a dictionary excluding all apis which satisfy a specific condition for instance key = 'HORIZONTAL' new dict would have every api that isnt a horizontal

  def unselectdict(self,dict2, idx, key):
      self.dict2 = dict2
      self.idx = idx
      self.key = key
      new_dict = dict((k, v) for k, v in self.dict2.items() if v[self.idx] != self.key)
      return new_dict

  # if key array = [0,none, ''] and various rows have this in a specific cell value for a specific row this api will be ignored

  def multiunselectdict(self,dict2, idx, keyarr):
      self.dict2 = dict2
      self.idx = idx
      self.keyarr = keyarr
      new_dict = dict((k, v) for k, v in self.dict2.items() if v[self.idx] not in self.keyarr)
      return new_dict

  def organize_data_and_calc(self):

    #import Analagous well Uncertainty Data and Offset Well Data from Well Inventory

    offset_wells_pre = self.offsetwells()

    # Import Planned Wells and Formation Tops
    #r_l = '/home/george/Downloads/ml for traders/rl/SFcalcs/Cedar Canyon 06 07_V3.2_Final Scenario_Approved Scenario_New.xlsm'
    planned_wells, formation_tops  = Plannedwells(self.plannedwellfile).make_planned_well_hash()

    form_list = sorted(formation_tops.values())
    tvd_deep_for_inc_only = form_list[0] - 200
    #print tvd_deep_for_inc_only
    tvd_deep_for_no_surv = form_list[0] - 1000
    #print tvd_deep_for_no_surv

    offset_wells_pre_2 = OrderedDict()
    for k, v in offset_wells_pre.items():
        if (v[3] >= tvd_deep_for_inc_only and ('INC ONLY' in v[4])) or (v[3] >= tvd_deep_for_no_surv and ('NO' in v[4])):
            offset_wells_pre_2[k] = v
    #print offset_wells_pre_2

    offset_laterals_pre = self.selectdict(offset_wells_pre_2,idx =5 ,key = 'HORIZONTAL' )
    #print len(offset_laterals_pre.keys())
    offset_directional = self.selectdict(offset_wells_pre_2,idx =5 ,key = 'DIRECTIONAL' )

    offset_wells_hash = offset_wells_pre_2

    offset_laterals = offset_laterals_pre


    # Well Calculations Case 1 ( Deep Well and  planned lateral Well)
    # ------------------------------------------------------------------ -------------------------------------------------------------------------------------------------
    api = list(offset_wells_hash.keys()) #3
    wellnames = list(planned_wells.keys())
    wells = []
    select_api = []
    surv = []
    select_x = []
    select_y = []
    select_tvd = []
    sf_one = []
    sf_two = []
    ctr_to_ctr = []
    stat = []
    gyro = []
    pr = []
    lat = []
    wells_1 = []
    select_api_1 = []
    surv_1 = []
    select_x_1 = []
    select_y_1 = []
    select_tvd_1 = []
    sf_one_1 = []
    sf_two_1 = []
    ctr_to_ctr_1 = []
    stat_1 = []
    gyro = []
    pr_1 = []
    lat_1 = []
    case = []
    case_1 = []
    #print formation_tops
    #print planned_wells
    for el1 in wellnames:
        toe_x = planned_wells[el1][2]
        toe_y = planned_wells[el1][3]
        heel_y = planned_wells[el1][1]
        heel_x = planned_wells[el1][0]
        y_shl = planned_wells[el1][4]
        kop_y = planned_wells[el1][5]
        build = planned_wells[el1][6]
        x_shl = planned_wells[el1][7]
        kop_x = planned_wells[el1][8]
        direction_planned_well = planned_wells[el1][9]
    # N_S Well Calculations-------------------------------------------------------------------------------------------------------------------------------------------------------------------
        if direction_planned_well == 'N-S':
            for k,v in formation_tops.items():
                if k in el1:
                    planned_well_tvd = v
                    break
            for el in api: #4
                status = offset_wells_hash[el][2]
                survey = offset_wells_hash[el][4]
                x_offset = offset_wells_hash[el][0]
                y_offset = offset_wells_hash[el][1]
                tvd_offset = offset_wells_hash[el][3]
                profile = offset_wells_hash[el][5]
                if build == '':
                    if (kop_y < y_offset and y_offset < heel_y) or (heel_y < y_offset and y_offset < kop_y):
                        lat_dist_fr_heel = 0
                    else:
                        lat_dist_fr_heel = abs(float(heel_y - y_offset))
                    lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                    if 'INC ONLY' in survey :
                        str1 = 'INC'
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( float((vert_uncert)**(2)) + float((lat_uncertainty)**(2)) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #5
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #6
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #7
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #8
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                        elif tvd_offset <= planned_well_tvd:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ((  (toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #9
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #10
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #11
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #12
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                    else:
                        str1 = ''
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #13
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #14
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #15
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})

                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #16
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                        elif  planned_well_tvd >= tvd_offset:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ( ((toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #17
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #18
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                            else:
                                ctr_to_ctr_dist = float( ( ((heel_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #19
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #20
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})



                else:
                    if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                        lat_dist_fr_heel = 0
                    else :
                        lat_dist_fr_heel = abs(float(heel_y - y_offset))
                    lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                    if  'INC ONLY' in survey:
                        str1 = 'INC'
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #21
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #22
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #23
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #24
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                        elif tvd_offset <= planned_well_tvd:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ((  (toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #25
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #26
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #27
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #28
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                    else :
                        str1 = ''
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #29
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #30
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_x - x_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #31
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})

                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #32
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                        elif  planned_well_tvd >= tvd_offset:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ( ((toe_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #33
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #34
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                            else:
                                ctr_to_ctr_dist = float( ( ((heel_x - x_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el) #35
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el) #36
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
    # E_W Calculations -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        elif direction_planned_well == 'E-W':
            for k,v in formation_tops.items():
                if k in el1:
                    planned_well_tvd = v
                    break
            for el in api:
                status = offset_wells_hash[el][2]
                survey = offset_wells_hash[el][4]
                x_offset = offset_wells_hash[el][0]
                y_offset = offset_wells_hash[el][1]
                tvd_offset = offset_wells_hash[el][3]
                profile = offset_wells_hash[el][5]
                if build == '':
                    if (kop_x < x_offset and x_offset < heel_x) or (heel_x < x_offset and x_offset < kop_x):
                        lat_dist_fr_heel = 0
                    else:
                        lat_dist_fr_heel = abs(float(heel_x - x_offset))
                    lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                    if  'INC ONLY' in survey:
                        str1 = 'INC'
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                        elif tvd_offset <= planned_well_tvd:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ((  (toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                    else:
                        str1 = ''
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd, str1)
                            one_sf = float(
                                (((vert_uncert)**(2) + (lat_uncertainty)**(2))**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})

                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                        elif  planned_well_tvd >= tvd_offset:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ( ((toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                            else:
                                ctr_to_ctr_dist = float( ( ((heel_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})



                else:
                    if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                        lat_dist_fr_heel = 0
                    else :
                        lat_dist_fr_heel = abs(float(heel_y - y_offset))
                    lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = 'case 1')
                    if  'INC ONLY' in survey:
                        str1 = 'INC'
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                        elif tvd_offset <= planned_well_tvd:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ((  (toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) + (tvd_offset- planned_well_tvd)**(2)  ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                    else :
                        str1 = ''
                        vert_uncert = self.vert_interp(tvd_offset,str1)
                        one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                        two_sf = float(one_sf*2)
                        if tvd_offset >= planned_well_tvd:
                            vert_uncert = self.vert_interp(planned_well_tvd,str1)
                            one_sf  = float((( (vert_uncert)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                            two_sf = float(one_sf*2)
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( (( (toe_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                            else:
                                ctr_to_ctr_dist = float( (( (heel_y - y_offset)**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})

                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


                        elif  planned_well_tvd >= tvd_offset:
                            if lat_dist_fr_heel >= 5000:
                                ctr_to_ctr_dist = float( ( ((toe_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})

                            else:
                                ctr_to_ctr_dist = float( ( ((heel_y - y_offset)**(2) + (tvd_offset- planned_well_tvd)**2 )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(1)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr,'Profile': pr, 'Lateral Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(1)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})



    # Well Calculations Case 2 ( 2 orthogonal laterals)
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
    api = list(offset_laterals.keys())
    wellnames = list(planned_wells.keys())
    for el1 in wellnames:
        toe_x = planned_wells[el1][2]
        toe_y = planned_wells[el1][3]
        heel_y = planned_wells[el1][1]
        heel_x = planned_wells[el1][0]
        y_shl = planned_wells[el1][4]
        kop_y = planned_wells[el1][5]
        build = planned_wells[el1][6]
        x_shl = planned_wells[el1][7]
        kop_x = planned_wells[el1][8]
        direction_planned_well = planned_wells[el1][9]
        #NS Planned well Calculations -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        if direction_planned_well == 'N-S':
            for k,v in formation_tops.items():
                if k in el1:
                    planned_well_tvd = v
                    break
            for el in api:
                status = offset_laterals[el][2]
                survey = offset_laterals[el][4]
                x_offset = offset_laterals[el][0]
                y_offset = offset_laterals[el][1]
                tvd_offset = offset_laterals[el][3]
                profile = offset_laterals[el][5]
                direction = offset_laterals[el][6]
                x_f = offset_laterals[el][7]
                y_f = offset_laterals[el][8]
                if (x_offset < heel_x and heel_x < x_f) or ( x_f < heel_x and heel_x < x_offset):
                    if direction == 'E-W':
                        if build == '':
                            if (kop_y < y_offset and  y_offset < toe_y) or (toe_y < y_offset and y_offset < kop_y): # basically saying if your within the y domain of the planned well
                                if (kop_y < y_offset and y_offset < heel_y) or (heel_y < y_offset and y_offset < kop_y):
                                    lat_dist_fr_heel = 0
                                else:
                                    lat_dist_fr_heel = abs(float(y_offset - heel_y))
                                lat_dist_intersect = abs(float(heel_x - x_offset))
                                lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = self.lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                two_sf = float(one_sf*2)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'Lat Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                        else:
                            if (y_shl < y_offset and y_offset < toe_y) or (toe_y < y_offset and y_offset < y_shl): # basically saying if you meet front build case and your not in its ydomain  dont do any calcs
                                if (heel_y < y_offset and y_offset < y_shl) or (y_shl < y_offset and y_offset < heel_y):
                                    lat_dist_fr_heel = 0
                                else :
                                    lat_dist_fr_heel = abs(float(heel_y - y_offset))
                                lat_dist_intersect = abs(float(heel_x - x_offset))
                                lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = self.lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                two_sf = float(one_sf*2)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'Lat Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
        # E_W Planned well Calculations -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        elif direction_planned_well == 'E-W':
            for k,v in formation_tops.items():
                if k in el1:
                    planned_well_tvd = v
                    break
            for el in api:
                status = offset_laterals[el][2]
                survey = offset_laterals[el][4]
                x_offset = offset_laterals[el][0]
                y_offset = offset_laterals[el][1]
                tvd_offset = offset_laterals[el][3]
                profile = offset_laterals[el][5]
                direction = offset_laterals[el][6]
                x_f = offset_laterals[el][7]
                y_f = offset_laterals[el][8]
                if (y_offset < heel_y and heel_y < y_f) or ( y_f < heel_y and heel_y < y_offset):
                    if direction == 'N-S':
                        if build == '':
                            if (kop_x < x_offset and  x_offset < toe_x) or (toe_x < x_offset and x_offset < kop_x): # basically saying if your within the y domain of the planned well
                                if (kop_x < x_offset and x_offset < heel_x) or (heel_x < x_offset and x_offset < kop_x):
                                    lat_dist_fr_heel = 0
                                else:
                                    lat_dist_fr_heel = abs(float(x_offset - heel_x))
                                lat_dist_intersect = abs(float(heel_y - y_offset))
                                lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = self.lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                two_sf = float(one_sf*2)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'Lat Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})
                        else:
                            if (x_shl < x_offset and x_offset < toe_x) or (toe_x < x_offset and x_offset < x_shl): # basically saying if you meet front build case and are within planned wells domain
                                if (heel_x < x_offset and x_offset < x_shl) or (x_shl < x_offset and x_offset < heel_x): # if intersect is between shl and heel lat dist from heel is 0 meaning still in build
                                    lat_dist_fr_heel = 0
                                else :
                                    lat_dist_fr_heel = abs(float(heel_x - x_offset))
                                lat_dist_intersect = abs(float(heel_y - y_offset))
                                lat_uncertainty = self.lat_interp(lat_dist_fr_heel, direction_planned_well, met = '')
                                lat_uncertainty1 = self.lat_interp(lat_dist_intersect, direction_planned_well, met = '')
                                one_sf  = float((( (lat_uncertainty1)**(2) + (lat_uncertainty)**(2) )**(0.5)))
                                two_sf = float(one_sf*2)
                                ctr_to_ctr_dist = float( (( abs((tvd_offset - planned_well_tvd))**(2) )**(0.5) ))
                                if two_sf >=  ctr_to_ctr_dist:
                                    need_gyro = 'True'
                                    wells.append(el1)
                                    select_api.append(el)
                                    surv.append(survey)
                                    stat.append(status)
                                    select_x.append(x_offset)
                                    select_y.append(y_offset)
                                    select_tvd.append(tvd_offset)
                                    sf_one.append(one_sf)
                                    sf_two.append(two_sf)
                                    ctr_to_ctr.append(ctr_to_ctr_dist)
                                    pr.append(profile)
                                    lat.append(lat_dist_fr_heel)
                                    case.append(2)
                                    anti_collision_sf = pd.DataFrame({'Well': wells, 'API': select_api, 'Survey': surv, 'X': select_x, 'Y': select_y, 'TVD': select_tvd, '1 SF': sf_one, '2 SF': sf_two, 'Center to Center Distance': ctr_to_ctr, 'Profile': pr, 'Lat Distance': lat,'case':case})
                                else:
                                    wells_1.append(el1)
                                    select_api_1.append(el)
                                    surv_1.append(survey)
                                    stat_1.append(status)
                                    select_x_1.append(x_offset)
                                    select_y_1.append(y_offset)
                                    select_tvd_1.append(tvd_offset)
                                    sf_one_1.append(one_sf)
                                    sf_two_1.append(two_sf)
                                    ctr_to_ctr_1.append(ctr_to_ctr_dist)
                                    pr_1.append(profile)
                                    lat_1.append(lat_dist_fr_heel)
                                    case_1.append(2)
                                    anti_collision_sf_2 = pd.DataFrame({'Well': wells_1, 'API': select_api_1, 'Survey': surv_1, 'X': select_x_1, 'Y': select_y_1, 'TVD': select_tvd_1, '1 SF': sf_one_1, '2 SF': sf_two_1, 'Center to Center Distance': ctr_to_ctr_1,'Profile': pr_1, 'Lateral Distance': lat_1,'case':case_1})


    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Labels
    #--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    #create labels


    label_hash_2 = OrderedDict()
    label_hash_new = OrderedDict()
    for idx,el in enumerate(select_api):
        #print idx
        #print el
        label_hash_new[select_api[idx]] = [select_tvd[idx], ctr_to_ctr[idx], sf_one[idx], sf_two[idx], pr[idx],surv[idx], wells[idx], select_x[idx], select_y[idx]]

    #label_arr = []
    for k,v in label_hash_new.items():
        if v[4] =='VERTICAL':
            #label_arr.append('Deep Vertical-Active' + ' ' + 'TVD=' + str(v[0]) + ' '+'API# =' + str(k) +str(v[5]) + ' available on NMOCD' + ' ' +str(v[1]) + 'from offset' +str(v[6]) + ' ' + '1.0 SF = '+ str(v[2]) +'ft,' + ' 2.0 SF = ' +str(v[3]) +'ft' + ' *below per Drilling A/C Guideline')
            label_hash_2[k] = [('Deep Vertical-Active' + ' ' + 'TVD=' + str(v[0]) + ' '+'API# =' +str(k) + ' ' +str(v[5]) + ' '+ 'available on NMOCD' + ' ' +str(v[1]) +' '+ ' from offset ' +str(v[6]) + ' ' + '1.0 SF = '+ str(v[2]) +'ft,' + ' 2.0 SF = ' +str(v[3]) +'ft' + ' *below per Drilling A/C Guideline'),v[7], v[8]]
        elif v[4] == 'HORIZONTAL':
            #label_arr.append('Lateral-Active' + ' ' + 'TVD=' + str(v[0]) + ' ' + 'API# =' + str(k) + str(v[5]) + ' available on NMOCD' + ' ' + str(v[1]) + 'from offset' + str(v[6]) + ' ' + '1.0 SF = ' + str(v[2]) + 'ft,' + ' 2.0 SF = ' + str(v[3]) + 'ft' + ' *below per Drilling A/C Guideline')
            label_hash_2[k] = [('Lateral-Active' + ' ' + 'TVD=' + str(v[0]) + ' ' + 'API# =' +str(k) + ' ' + str(v[5]) +' ' + 'available on NMOCD' + ' ' + str(v[1]) +' '+ ' from offset ' + str(v[6]) + ' ' + '1.0 SF = ' + str(v[2]) + 'ft,' + ' 2.0 SF = ' + str(v[3]) + 'ft' + ' *below per Drilling A/C Guideline'), v[7], v[8]]

    offset_wells_pre = self.offsetwells()
    flagged_arr = label_hash_2.keys()
    offset_wells_label_hash = OrderedDict()
    for k, v in offset_wells_pre.items():
        if k not in flagged_arr:
            offset_wells_label_hash[k] = [str(v[5]) + '-Active' + ' '+ 'API=' + str(k) + ' ' + 'TVD=' + str(v[3]), v[0], v[1]]

    offset_df = pd.DataFrame.from_dict(offset_wells_label_hash, orient='index',columns=['String', 'X','Y'])

    flagged_df = pd.DataFrame.from_dict(label_hash_2, orient='index',columns=['String', 'X','Y'])

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Writing DF  to .Xlsx
    #--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    writer = pd.ExcelWriter(self.outputfilename,engine='xlsxwriter')

    anti_collision_sf.to_excel(writer, sheet_name='Flagged Wells')
    anti_collision_sf_2.to_excel(writer, sheet_name='Wells not flagged')
    offset_df.to_excel(writer,sheet_name='Offset Labels')
    flagged_df.to_excel(writer,sheet_name='Flagged Labels')

    writer.save()

if __name__ =="__main__":

  # only 3 things need to be changed

  #1 change output file_name
  out_put_file_name = 'test.xlsx'
  #2  Well Inventory directory
  wellsfile = 'xx'
  #3 Planned wells directory
  planned_well_file_location = '/home/george/Downloads/ml for traders/rl/SFcalcs/Cedar Canyon 06 07_V3.2_Final Scenario_Approved Scenario_New.xlsm'
  l = len(planned_well_file_location.split('/'))
  path = ('/').join(planned_well_file_location.split('/')[:(l-1)])
  os.chdir(path)
  Sfcalcnew(wellsfile,planned_well_file_location,out_put_file_name).organize_data_and_calc()
