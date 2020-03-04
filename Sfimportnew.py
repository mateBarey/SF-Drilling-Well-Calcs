#-------------------------------------------------------------------------------
# Name:        Sfimportnew
# Purpose:
#
# Author:      George Raul Cubas
#
# Created:     X/XX/XXXX
# Copyright:   (c) George Raul Cubas
# Licence:     <MIT>
#-------------------------------------------------------------------------------

#**************************************************************************************
# System Setup
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
#**************************************************************************************


class Plannedwells:
  def __init__(self,wellfile):
    self.wellfile = wellfile

  def make_planned_well_hash(self):
    tolerance=5
    workbook = openpyxl.load_workbook(self.wellfile, data_only=True)
    sheet3 = workbook['Targets']
    sheet1 = workbook['Lease information']
    planned_wells = {}
    i = 0
    tot_wells = 0
    for row in range(65, sheet3.max_row +1):
        if sheet3['A' + str(row)].value == 1:
            tot_wells += 1

    for row in range(65, sheet3.max_row +1):
        if sheet3['A' + str(row)].value == 1:
            str_name = sheet3['B'+ str(row)].value
            y_shl = sheet3['F'+ str(row)].value
            x_shl = sheet3['E'+ str(row)].value
            kop_y = sheet3['F'+ str(row +1)].value
            kop_x = sheet3['E'+ str(row +1)].value
            az = sheet3['T'+ str(row)].value
            Well_name = str_name.split('#')[0][3:]+ str_name.split('#')[1][1]
            if sheet3['D' + str(row + 2)].value == 'Heel':
                heel_x = sheet3['E' + str(row + 2)].value
                heel_y = sheet3['F' + str(row + 2)].value
            if sheet3['D' + str(row + 3)].value == 'Toe':
                toe_x = sheet3['E' + str(row + 3)].value
                toe_y = sheet3['F' + str(row + 3)].value
            #build analysis
            # ---------------------------------------------------------------------------------------------------------------------------
            #note (remember for simplistic reasons az on sheet is not true az relative to grid north or mag north it simply is a geometric degree)
            #ns build conditions
            # ---------------------------------------------------------------------------------------------------------------------------
            if ( 90-tolerance <= az  and  az <= 90+tolerance )  or ( -90-tolerance <= az and az <= -90+tolerance):
                direction = 'N-S'
                if az <= (-88) and y_shl > kop_y:
                    build = 'frontbuild north to south'
                elif az >= (88) and kop_y > y_shl:
                    build = 'frontbuild south to north'
                else:
                    build = ''
            #ew build conditions
            # ---------------------------------------------------------------------------------------------------------------------------
            elif (0-tolerance <= az and az <= 3+tolerance) or (180-tolerance <= az and az <= 180+tolerance) or (-180-tolerance <= az and az <= -180+tolerance):
                direction = 'E-W'
                if  (0-tolerance <= az and az <= 3+tolerance) and x_shl < kop_x:
                    build = 'frontbuild west to east'
                elif (180-tolerance <= az and az <= 180+tolerance) and x_shl > kop_x:
                    build = 'frontbuild east to west'
                else:
                    build = ''

            planned_wells[Well_name] = [heel_x, heel_y, toe_x, toe_y, y_shl, kop_y, build, x_shl, kop_x, direction]

            i += 1

            if i > tot_wells:
                break
            # ---------------------------------------------------------------------------------------------------------------------------
    elevation = sheet1['H14'].value
    formation_tops = {}
    for row in range(27, sheet1.max_row +1):
        bench_tops = sheet1['G'+ str(row)].value
        tvd_tops = sheet1['I' + str(row)].value
        formation_tops[bench_tops] = tvd_tops
        if (sheet1['G'+ str(row + 1)].value) == '' or (sheet1['G'+ str(row + 1)].value) == None:
            break
    for k,v in formation_tops.items():
        if v == elevation:
            del formation_tops[k]

    return planned_wells, formation_tops


